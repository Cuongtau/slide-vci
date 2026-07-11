import sys, io, os, json, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

rows = json.load(open(r"d:\SourceCode\Slide VCI\_knl_master.json", encoding="utf-8"))

# --- fixes ---
for r in rows:
    if r["name"].upper().replace("-"," ") == "NGUYEN VAN VUONG":
        r["name"] = "Nguyễn Văn Vương"
    if "Khúc Vân Anh" in r["name"]:
        r["mnv"] = "VCI.0033"

# --- remove Phúc ---
rows = [r for r in rows if "Phúc" not in r["name"]]

TEAM_LABEL = {"BA":"BA","DEV":"DEV","QC":"QA/QC","UIUX":"UI/UX"}
catkeys = ["co_ban","chuyen_mon","mem","quan_ly","vh_dd","cong_hien"]

def rnd(x,d=1):
    return None if x is None else round(float(x),d)

# --- build Vương combined (2 tab) ---
vuong = [r for r in rows if r["name"]=="Nguyễn Văn Vương"]
vuong_combined = None
if len(vuong) == 2:
    def avg2(getter):
        xs=[getter(r) for r in vuong]; xs=[x for x in xs if isinstance(x,(int,float))]
        return round(sum(xs)/len(xs),2) if xs else None
    vuong_combined = {
        "mnv":"", "name":"Nguyễn Văn Vương", "role":"BE+FE (TB 2 tab)", "team":"DEV",
        "cats":{k:avg2(lambda r,k=k:r["cats"].get(k)) for k in catkeys},
        "tot":{"ns":avg2(lambda r:r["tot"]["ns"]),
               "cbqltt":avg2(lambda r:r["tot"]["cbqltt"]),
               "capt":avg2(lambda r:r["tot"]["capt"])},
        "combined":True,
    }

# ================= Excel =================
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Tổng hợp KNL"
blue   = PatternFill("solid", fgColor="B4C7E7"); blue2 = PatternFill("solid", fgColor="8EAADB")
orange = PatternFill("solid", fgColor="F4B183");  teamf = PatternFill("solid", fgColor="FFE699")
tbf    = PatternFill("solid", fgColor="E2EFDA");   grey  = PatternFill("solid", fgColor="D9D9D9")
leadf  = PatternFill("solid", fgColor="FFF2CC")
white  = Font(bold=True, color="1F3864")
thin=Side(style="thin",color="BFBFBF"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
center=Alignment(horizontal="center",vertical="center",wrap_text=True)
left=Alignment(horizontal="left",vertical="center",wrap_text=True)

headers=["MNV","Họ tên","Vị trí","Bộ phận","Năng lực cơ bản","Năng lực chuyên môn","Năng lực mềm",
         "Năng lực quản lý","Văn hóa - Đạo đức","Cống hiến","Tổng điểm NS","Tổng CBQLTT","Tổng QL trên 1 cấp"]
NC=len(headers)

ws.cell(1,1,"BẢNG TỔNG HỢP ĐÁNH GIÁ KHUNG NĂNG LỰC 2026 — ĐỘI CĐS (PM Bùi Xuân Cương)")
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=NC); ws.cell(1,1).font=Font(bold=True,size=13,color="1F3864"); ws.cell(1,1).alignment=center

ws.cell(2,5,"NHÂN VIÊN (tự đánh giá)"); ws.merge_cells(start_row=2,start_column=5,end_row=2,end_column=11)
ws.cell(2,5).fill=blue2; ws.cell(2,5).font=Font(bold=True,color="FFFFFF"); ws.cell(2,5).alignment=center
for c in range(1,5): ws.cell(2,c).fill=grey
ws.cell(2,12,"HỘI ĐỒNG"); ws.merge_cells(start_row=2,start_column=12,end_row=2,end_column=13)
ws.cell(2,12).fill=orange; ws.cell(2,12).font=Font(bold=True,color="7F3300"); ws.cell(2,12).alignment=center

HR=3
for i,h in enumerate(headers,1):
    cell=ws.cell(HR,i,h); cell.font=white; cell.alignment=center; cell.border=border
    cell.fill = blue if 5<=i<=11 else (orange if i>=12 else grey)

def write_row(rr, r, fill=None, bold_name=False):
    c=r["cats"]; t=r["tot"]
    vals=[r["mnv"],r["name"],r["role"],TEAM_LABEL.get(r["team"],r["team"])]
    vals+=[rnd(c.get(k),2) for k in catkeys]
    vals+=[rnd(t["ns"],2),rnd(t["cbqltt"],2),rnd(t["capt"],2)]
    for i,v in enumerate(vals,1):
        cell=ws.cell(rr,i,v); cell.border=border
        cell.alignment=left if i==2 else center
        if fill: cell.fill=fill
        if i==11: cell.font=Font(bold=True)
        if i==2 and bold_name: cell.font=Font(bold=True)

def team_avg(persons,key,sub=None,skip_default=False):
    xs=[]
    for r in persons:
        v = r["cats"].get(sub) if sub else r["tot"].get(key)
        if not isinstance(v,(int,float)): continue
        if skip_default and v>=99: continue   # bỏ ô mặc định template 99.75
        xs.append(v)
    return round(sum(xs)/len(xs),1) if xs else None

order=["BA","DEV","QC","UIUX"]
rr=HR+1
for team in order:
    grp=[r for r in rows if r["team"]==team]
    if not grp: continue
    # persons list for averaging (Vương -> combined once)
    if team=="DEV" and vuong_combined:
        persons=[r for r in grp if r["name"]!="Nguyễn Văn Vương"]+[vuong_combined]
        # display: Vương 2 tab + combined first (Lead), then others by NS desc
        others=sorted([r for r in grp if r["name"]!="Nguyễn Văn Vương"],
                      key=lambda r:-(r["tot"]["ns"] or 0))
        vb=[r for r in vuong if r["role"]=="Dev BE"]; vf=[r for r in vuong if r["role"]=="Dev FE"]
        display=vb+vf+[vuong_combined]+others
    else:
        persons=grp
        display=sorted(grp,key=lambda r:-(r["tot"]["ns"] or 0))
    for r in display:
        if r.get("combined"):
            write_row(rr,r,fill=leadf,bold_name=True); rr+=1
        else:
            write_row(rr,r); rr+=1
    # TB row
    npers=len(persons)
    st=["",f"TRUNG BÌNH {TEAM_LABEL.get(team,team)} ({npers} NS)","",""]
    st+=[team_avg(persons,None,k) for k in catkeys]
    st+=[team_avg(persons,"ns"),team_avg(persons,"cbqltt",skip_default=True),
         team_avg(persons,"capt",skip_default=True)]
    for i,v in enumerate(st,1):
        cell=ws.cell(rr,i,v); cell.fill=tbf; cell.border=border
        cell.font=Font(bold=True,italic=True); cell.alignment=center if i!=2 else left
    rr+=1

rr+=1
note=("Ghi chú: (1) Đã LOẠI nhân sự Thái Doãn Phúc theo yêu cầu.  "
      "(2) Nguyễn Văn Vương có 2 tab BE & FE — hiển thị cả 2 dòng + dòng 'TB 2 tab' (nền vàng); TB team DEV chỉ tính Vương 1 lần.  "
      "(3) 6 cột năng lực = cột 'Điểm quy đổi /100' của từng năng lực (đã nhân trọng số) — TỔNG 6 cột = Tổng điểm NS. Trọng số: CM 50%, các mục khác 10%.  "
      "(4) Ô CBQLTT/QL trên còn trống = chưa chấm; giá trị 99.75 = mặc định template (đã loại khỏi dòng trung bình).")
ws.cell(rr,1,note); ws.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=NC)
ws.cell(rr,1).alignment=Alignment(wrap_text=True,vertical="top"); ws.cell(rr,1).font=Font(italic=True,size=9,color="808080")
ws.row_dimensions[rr].height=60

for i,w in enumerate([10,24,15,9,9,10,8,9,10,8,9,9,11],1):
    ws.column_dimensions[get_column_letter(i)].width=w
ws.row_dimensions[HR].height=42
ws.freeze_panes="E4"; ws.auto_filter.ref=f"A{HR}:{get_column_letter(NC)}{HR}"

out=r"d:\VCI Group\CÔNG TY CỔ PHẦN CÔNG NGHỆ SỐ VCI\Phạm Thị Ngọc (BO-HCNS) - KNL_CĐS_PM_Bùi Xuân Cương\00_TongHop_DanhGia_KNL_DoiCDS_Cuong.xlsx"
try:
    wb.save(out); print("SAVED:",out)
except Exception as e:
    alt=r"C:\Users\ASUS\AppData\Local\Temp\claude\d--VCI-Group-C-NG-TY-C--PH-N-C-NG-NGH--S--VCI-Ph-m-Th--Ng-c--BO-HCNS----KNL-C-S-PM-B-i-Xu-n-C--ng\ba27c969-8872-4913-beff-6577f8791ddc\scratchpad\00_TongHop_DanhGia_KNL_DoiCDS_Cuong.xlsx"
    wb.save(alt); print("SAVED(scratch):",alt,"| lý do:",e)

# also print combined table for chat
print("\n--- VƯƠNG ---")
for r in vuong+([vuong_combined] if vuong_combined else []):
    c=r["cats"]; t=r["tot"]
    print(r["role"], [c.get(k) for k in catkeys], "NS=",t["ns"],"CBQLTT=",t["cbqltt"],"capt=",t["capt"])
# team averages
for team in order:
    grp=[r for r in rows if r["team"]==team]
    if team=="DEV" and vuong_combined:
        persons=[r for r in grp if r["name"]!="Nguyễn Văn Vương"]+[vuong_combined]
    else: persons=grp
    print(f"TB {team} ({len(persons)}): NS={team_avg(persons,'ns')} cats="+
          str([team_avg(persons,None,k) for k in catkeys]))
