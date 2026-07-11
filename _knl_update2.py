import sys, io, os, json, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

DST = r"C:\Users\ASUS\AppData\Local\Temp\claude\d--VCI-Group-C-NG-TY-C--PH-N-C-NG-NGH--S--VCI-Ph-m-Th--Ng-c--BO-HCNS----KNL-C-S-PM-B-i-Xu-n-C--ng\ba27c969-8872-4913-beff-6577f8791ddc\scratchpad\knl"
SKIP=("truyền thông","khảo sát","param","hướng dẫn")
CAT_KEYS=[("co_ban","năng lực cơ bản"),("chuyen_mon","năng lực chuyên môn"),
          ("mem","năng lực mềm"),("quan_ly","năng lực quản lý"),
          ("vh_dd","văn ho"),("cong_hien","cống hiến")]
ROLE_MAP={"ba":"BA","dev be":"Dev BE","dev fe":"Dev FE","dev mobile flutter":"Mobile Dev","qc":"QC","uiux":"UI/UX"}

def find_label(ws,label,maxr=12,maxc=20):
    label=label.lower()
    for r in range(1,min(maxr,ws.max_row)+1):
        for c in range(1,min(maxc,ws.max_column)+1):
            v=ws.cell(r,c).value
            if v and label in str(v).lower():
                for cc in range(c+1,min(c+6,ws.max_column)+1):
                    vv=ws.cell(r,cc).value
                    if vv not in (None,""): return str(vv).strip()
    return ""
def is_eval(ws):
    for r in range(1,min(ws.max_row,260)+1):
        v=ws.cell(r,2).value
        if v and "tổng điểm" in str(v).lower(): return True
    return False
def gnum(ws,r,c):
    v=ws.cell(r,c).value
    try: return round(float(v),2)
    except: return None

def extract(path, team, fallname):
    wb=openpyxl.load_workbook(path,data_only=True)
    out=[]
    for ws in wb.worksheets:
        if any(s in ws.title.lower() for s in SKIP): continue
        if not is_eval(ws): continue
        nm=find_label(ws,"họ và tên")
        if not nm or nm.lower().startswith("loại"): nm=fallname
        role=ROLE_MAP.get(ws.title.strip().lower(),ws.title.strip())
        cats={}; tot={"ns":None,"cbqltt":None,"capt":None}
        for r in range(1,min(ws.max_row,260)+1):
            b=ws.cell(r,2).value; c=ws.cell(r,3).value
            low=((str(b) if b else "")+" "+(str(c) if c else "")).strip().lower()
            if low.startswith("tổng điểm"):
                tot={"ns":gnum(ws,r,13),"cbqltt":gnum(ws,r,15),"capt":gnum(ws,r,17)}
            ct=str(c).strip().lower() if c else ""
            for key,kw in CAT_KEYS:
                if ct.startswith(kw) and key not in cats:
                    M=gnum(ws,r,13); L=gnum(ws,r,12)
                    cats[key]=round(M*10/L,1) if (M is not None and L) else None
        out.append({"mnv":"","name":nm,"role":role,"cats":cats,"tot":tot,"team":team})
    return out

new=[]
new += extract(os.path.join(DST,"vuong_new.xlsx"),"DEV","Nguyễn Văn Vương")
new += extract(os.path.join(DST,"linh_new.xlsx"),"BA","Nguyễn Thị Thùy Linh")
for r in new:
    if r["name"].upper().replace("-"," ")=="NGUYEN VAN VUONG": r["name"]="Nguyễn Văn Vương"

print("=== DỮ LIỆU MỚI ===")
ck=["co_ban","chuyen_mon","mem","quan_ly","vh_dd","cong_hien"]
for r in new:
    print(f"{r['name']:22} {r['role']:8} cats={[r['cats'].get(k) for k in ck]} "
          f"NS={r['tot']['ns']} CBQLTT={r['tot']['cbqltt']} capt={r['tot']['capt']}")

# --- update master json ---
master=json.load(open(r"d:\SourceCode\Slide VCI\_knl_master.json",encoding="utf-8"))
def key(r): return (r["name"], r["role"])
newkeys={key(r) for r in new}
# names being replaced
repl_names={"Nguyễn Văn Vương","Nguyễn Thị Thùy Linh"}
master=[r for r in master if r["name"] not in repl_names]
# also drop the plain filename-derived Vuong ('NGUYEN-VAN-VUONG') if any
master=[r for r in master if r["name"].upper().replace("-"," ")!="NGUYEN VAN VUONG"]
master += new
json.dump(master,open(r"d:\SourceCode\Slide VCI\_knl_master.json","w",encoding="utf-8"),ensure_ascii=False,indent=1)
print("\nMaster rows now:",len(master))
