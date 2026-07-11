import sys, io, os, re, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.utils import get_column_letter

DST = r"C:\Users\ASUS\AppData\Local\Temp\claude\d--VCI-Group-C-NG-TY-C--PH-N-C-NG-NGH--S--VCI-Ph-m-Th--Ng-c--BO-HCNS----KNL-C-S-PM-B-i-Xu-n-C--ng\ba27c969-8872-4913-beff-6577f8791ddc\scratchpad\knl"
man = json.load(open(os.path.join(DST,"manifest.json"),encoding="utf-8-sig"))
if isinstance(man,dict): man=[man]

SKIP=("truyền thông","khảo sát","param","hướng dẫn")
CAT_KEYS=[("co_ban","năng lực cơ bản"),("chuyen_mon","năng lực chuyên môn"),
          ("mem","năng lực mềm"),("quan_ly","năng lực quản lý"),
          ("vh_dd","văn ho"),("cong_hien","cống hiến")]
ROLE_MAP={"ba":"BA","dev be":"Dev BE","dev fe":"Dev FE","dev mobile flutter":"Mobile Dev",
          "qc":"QC","uiux":"UI/UX"}

def find_label(ws,label,maxr=12,maxc=20):
    label=label.lower()
    for r in range(1,min(maxr,ws.max_row)+1):
        for c in range(1,min(maxc,ws.max_column)+1):
            v=ws.cell(r,c).value
            if v and label in str(v).lower():
                for cc in range(c+1,min(c+6,ws.max_column)+1):
                    vv=ws.cell(r,cc).value
                    if vv not in (None,""): return str(vv).strip()
    return ""

def find_mnv(ws, orig):
    for lbl in ("mã nhân viên","mã nv","mnv","mã số"):
        v=find_label(ws,lbl)
        if v and not str(v).lower().startswith("họ"): return v
    m=re.search(r"VCI[.\s]?0*(\d{2,4})", orig)
    if m: return "VCI."+m.group(1)
    return ""

def is_eval(ws):
    for r in range(1,min(ws.max_row,260)+1):
        v=ws.cell(r,2).value
        if v and "tổng điểm" in str(v).lower(): return True
    return False

def gnum(ws,r,c):
    v=ws.cell(r,c).value
    try: return round(float(v),2)
    except: return None

def name_from_orig(orig):
    s=re.sub(r"\.xlsx$","",orig)
    s=re.sub(r"^(KNL_|DGNL |DGNL_)","",s)
    s=re.sub(r"^VCI\.QĐ\.3058_Hệ thống KNL_","",s)
    s=re.sub(r"^(D0?2_|Q0?1_)","",s)
    s=re.sub(r"^VCI\.\d+\.?\s*","",s)
    s=re.split(r"_",s)[0].strip()
    return s

def extract(ws, orig):
    nm=find_label(ws,"họ và tên")
    if not nm or nm.lower().startswith("loại"): nm=name_from_orig(orig)
    role=ROLE_MAP.get(ws.title.strip().lower(), ws.title.strip())
    cats={}; tot={"ns":None,"cbqltt":None,"capt":None}
    for r in range(1,min(ws.max_row,260)+1):
        b=ws.cell(r,2).value; c=ws.cell(r,3).value
        low=((str(b) if b else "")+" "+(str(c) if c else "")).strip().lower()
        if low.startswith("tổng điểm"):
            tot={"ns":gnum(ws,r,13),"cbqltt":gnum(ws,r,15),"capt":gnum(ws,r,17)}
        ct=str(c).strip().lower() if c else ""
        for key,kw in CAT_KEYS:
            if ct.startswith(kw) and key not in cats:
                M=gnum(ws,r,13); L=gnum(ws,r,12)
                # Lấy ĐÚNG cột "Điển quy đổi /100" (cột M) của từng năng lực -> tổng 6 cột = Tổng NS
                cats[key]=M
    return {"mnv":find_mnv(ws,orig),"name":nm,"role":role,"cats":cats,"tot":tot}

rows=[]
seen_files=set()
for e in man:
    seen_files.add(e["orig"])
    path=os.path.join(DST,e["ascii"])
    try:
        wb=openpyxl.load_workbook(path,data_only=True)
    except Exception as ex:
        print("LOADERR",e["orig"],ex); continue
    n_eval=0
    for ws in wb.worksheets:
        if any(s in ws.title.lower() for s in SKIP): continue
        if not is_eval(ws): continue
        n_eval+=1
        rec=extract(ws,e["orig"]); rec["team"]=e["team"]; rec["orig"]=e["orig"]
        rows.append(rec)
    print(f"OK {e['team']:5} {e['orig'][:45]:47} sheets_eval={n_eval}")

order={"BA":0,"DEV":1,"QC":2,"UIUX":3}
rows.sort(key=lambda r:(order.get(r["team"],9), -(r["tot"]["ns"] or 0)))
print("\nTOTAL EVAL ROWS:",len(rows))
json.dump(rows,open(r"d:\SourceCode\Slide VCI\_knl_master.json","w",encoding="utf-8"),ensure_ascii=False,indent=1)
